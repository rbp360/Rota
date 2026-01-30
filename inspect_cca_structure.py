
import openpyxl
import os

EXCEL_PATH = r"c:\Users\rob_b\Rota\temp_rota.xlsx"

def inspect_cca():
    if not os.path.exists(EXCEL_PATH):
        print(f"File not found: {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if "CCA" not in wb.sheetnames:
        print("CCA sheet not found")
        return

    sheet = wb["CCA"]
    print(f"Sheet: CCA")
    
    # Print first 20 rows and 15 columns
    for r_idx, row in enumerate(sheet.iter_rows(max_row=20, max_col=15, values_only=True)):
        print(f"Row {r_idx+1}: {row}")

if __name__ == "__main__":
    inspect_cca()
