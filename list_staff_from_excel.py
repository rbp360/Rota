
import openpyxl
import re

EXCEL_PATH = r"c:\Users\rob_b\Rota\temp_rota.xlsx"

def get_all_potential_staff():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    found_names = set()
    
    # 1. From Staff Sheet
    for name in wb.sheetnames:
        if name.lower() not in ['instructions', 'summary', 'record', 'absencerecord', 'sheet3', 'sheet1', 'cca', 'y56eal', 'cover']:
             found_names.add(name)

    # 2. From CCA Sheet (using our new logic)
    if "CCA" in wb.sheetnames:
        sheet = wb["CCA"]
        days_list = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        day_cols = {}
        header_row = None
        
        # Find headers
        for r_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
            row_str = [str(c).lower() if c else "" for c in row]
            found = False
            for d in days_list:
                if d.lower() in row_str:
                    found = True
                    # Map day -> staff col (Activity | Staff)
                    c_idx = row_str.index(d.lower())
                    day_cols[d] = c_idx + 1 # Staff is next column
            if found:
                header_row = r_idx + 1
                break
        
        if header_row:
             for row in sheet.iter_rows(min_row=header_row+1, values_only=True):
                 if not any(row): continue
                 for day, staff_col_idx in day_cols.items():
                     if staff_col_idx < len(row):
                         val = row[staff_col_idx]
                         if val:
                             # Split
                             parts = re.split(r'[+=\n&]', str(val))
                             for p in parts:
                                 p = p.strip()
                                 if p:
                                     # Basic clean
                                     if "(" in p: p = p.split("(")[0].strip()
                                     found_names.add(p)

    print("\n--- FOUND STAFF LIST ---")
    for n in sorted(list(found_names)):
        print(n)

if __name__ == "__main__":
    get_all_potential_staff()
