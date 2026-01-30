
import openpyxl
import os

EXCEL_PATH = r"c:\Users\rob_b\Rota\temp_rota.xlsx"

def debug_cca():
    if not os.path.exists(EXCEL_PATH):
        print(f"File not found: {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    if "CCA" not in wb.sheetnames:
        print("CCA sheet not found")
        return

    sheet = wb["CCA"]
    print("--- CCA SHEET HEADERS AND FIRST 5 ROWS ---")
    rows = list(sheet.iter_rows(max_row=10, values_only=True))
    for i, row in enumerate(rows):
        print(f"Row {i}: {row}")

if __name__ == "__main__":
    debug_cca()
