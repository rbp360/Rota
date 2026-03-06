import openpyxl
wb = openpyxl.load_workbook('GV cover and staff absence.xlsx', read_only=True, data_only=True)
with open('sheet_preview.txt', 'w', encoding='utf-8') as f:
    for name in wb.sheetnames:
        f.write(f"--- Sheet: {name} ---\n")
        sheet = wb[name]
        for row in sheet.iter_rows(max_row=5, max_col=5, values_only=True):
            f.write(str(row) + '\n')
        f.write('\n')
