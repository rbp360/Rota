import os
import re
import openpyxl
import sqlite3
import firebase_admin
from firebase_admin import credentials, firestore

# Configuration
EXCEL_PATH = "GV cover and staff absence.xlsx"
DB_PATH = "data_archive/rota.db"
CREDS_PATH = "rotaai-49847-d923810f254e.json"
DAYS_LIST = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

# Specialist list
SPECIALISTS = [
    "Daryl", "Becky", "Billy", "Jinny", "Ginny", "Ben", "Faye", 
    "Claire", "Jake", "Retno", "Jacinta", "Sunny", "Mr Ben", "Nick C"
]
FREE_SUBJECTS = ['thai', 'music', 'pe', 'p.e.', 'phse']

def clean_name(name):
    if not name: return ""
    n = str(name).strip()
    n = re.sub(r'^(mr|mrs|ms|miss|k\.|kun\s|k\s)\s*', '', n, flags=re.IGNORECASE).strip()
    n = n.split('(')[0].strip()
    return n

def init_firestore():
    if not os.path.exists(CREDS_PATH): return None
    try:
        if not firebase_admin._apps:
            cred = credentials.Certificate(CREDS_PATH)
            firebase_admin.initialize_app(cred)
        return firestore.client()
    except: return None

def run_targeted_fix():
    print(f"--- STARTING TARGETED FIX FOR PERIOD 1 ---")
    
    # 1. Load Excel
    print(f"Reading {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    
    # 2. Setup DBs
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    db_cloud = init_firestore()
    if not db_cloud:
        print("Warning: Could not initialize Cloud Firestore. Only local DB will be updated.")
    
    # 3. Get Staff
    cursor.execute("SELECT id, name FROM staff")
    staff_mems = cursor.fetchall()
    
    local_updates = 0
    cloud_updates = 0
    
    for s in staff_mems:
        staff_id_local = s['id']
        original_name = s['name']
        staff_name = clean_name(original_name)
        
        if not staff_name: continue
        
        # Find sheet
        target_sheet = None
        for sname in sheet_names:
            if staff_name.lower() in sname.lower() or sname.lower() in staff_name.lower():
                target_sheet = sname
                break
        
        if not target_sheet:
            if staff_name == "Claire" and "ME" in sheet_names: target_sheet = "ME"
            else: continue

        print(f"Processing: {original_name}")
        sheet = wb[target_sheet]
        
        day_cols = {}
        for row in sheet.iter_rows(max_row=10, values_only=True):
            for i, cell in enumerate(row):
                if cell:
                    c_val = str(cell).strip().lower()
                    for d in DAYS_LIST:
                        if d.lower() in c_val: day_cols[d] = i
            if day_cols: break
            
        if not day_cols: day_cols = {d: i+4 for i, d in enumerate(DAYS_LIST)}

        p1_row_data = None
        for row in sheet.iter_rows(max_row=60, values_only=True):
            for cell in row[:5]:
                if cell:
                    c_val = str(cell).strip().lower()
                    if re.search(r'(^|\b)(period\s*|p\.?\s*|)1(\b|$)', c_val):
                        p1_row_data = row
                        break
            if p1_row_data: break
        
        if p1_row_data:
            if db_cloud:
                # Find the Firestore matching staff doc
                # In this system, Firestore IDs are often the same as names or UUIDs.
                # Let's try to match by name in Firestore
                f_docs = db_cloud.collection("staff").where("name", "==", original_name).limit(1).get()
                f_staff_id = f_docs[0].id if f_docs else None
                if not f_staff_id:
                    # Fallback search
                    f_docs = db_cloud.collection("staff").get()
                    for doc in f_docs:
                        if clean_name(doc.to_dict().get("name", "")) == staff_name:
                            f_staff_id = doc.id
                            break
            else:
                f_staff_id = None

            for day, col_idx in day_cols.items():
                if col_idx < len(p1_row_data):
                    activity = str(p1_row_data[col_idx]).strip() if p1_row_data[col_idx] else ""
                    clean_act = activity.lower()
                    is_free = 0 # Default to not free

                    if not clean_act or any(x in clean_act for x in ['free', 'available', 'none', 'nan']):
                        is_free = 1
                    elif "assembly" in clean_act:
                        # Assembly is generally a free period for cover
                        is_free = 1
                    elif staff_name not in SPECIALISTS:
                        if any(sub in clean_act for sub in FREE_SUBJECTS):
                            is_free = 1
                    
                    # Update SQLite
                    cursor.execute("UPDATE schedules SET activity=?, is_free=? WHERE staff_id=? AND day_of_week=? AND period=1", 
                                 (activity, is_free, staff_id_local, day))
                    local_updates += 1
                    
                    # Update Cloud
                    if f_staff_id:
                        db_cloud.collection("staff").document(f_staff_id).collection("schedules").document(f"{day}_1").set({
                            "day_of_week": day, "period": 1, "activity": activity, "is_free": bool(is_free)
                        }, merge=True)
                        cloud_updates += 1
            
    conn.commit()
    conn.close()
    print(f"\n✅ FINISHED.")
    print(f"- Local SQLite: {local_updates} P1 records updated.")
    print(f"- Cloud Firestore: {cloud_updates} P1 records updated.")

if __name__ == "__main__":
    run_targeted_fix()
