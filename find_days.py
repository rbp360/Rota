import openpyxl

def find_days():
    wb = openpyxl.load_workbook(r"c:\Users\rob_b\Rota\GV cover and staff absence.xlsx", read_only=True, data_only=True)
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    
    # Check the first few staff sheets
    subset = [n for n in wb.sheetnames if n.lower() not in ['instructions', 'summary', 'record']][:3]
    
    for name in subset:
        print(f"--- Checking Sheet: {name} ---")
        sheet = wb[name]
        for row_idx, row in enumerate(sheet.iter_rows(max_row=10, max_col=20, values_only=True)):
            for col_idx, value in enumerate(row):
                if str(value).strip() in days:
                    print(f"Found {value} at Row {row_idx + 1}, Col {col_idx + 1}")

if __name__ == "__main__":
    find_days()
