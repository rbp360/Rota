import openpyxl

EXCEL_PATH = r"c:\Users\rob_b\Rota\GV cover and staff absence.xlsx"

def debug_staff(staff_name):
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    if staff_name not in wb.sheetnames:
        print(f"Sheet {staff_name} not found")
        return
    
    sheet = wb[staff_name]
    print(f"\n--- DEBUGGING {staff_name} ---")
    for r_idx, row in enumerate(sheet.iter_rows(max_row=12, max_col=15, values_only=True)):
        print(f"Row {r_idx + 1}: {row}")

if __name__ == "__main__":
    debug_staff("Ben")
    debug_staff("Amanda")
