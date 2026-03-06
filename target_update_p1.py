import os
import re
import openpyxl
import firebase_admin
from firebase_admin import credentials, firestore

# Configuration
EXCEL_PATH = "GV cover and staff absence.xlsx"
CREDS_PATH = "rotaai-49847-d923810f254e.json"
DAYS_LIST = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

# Specialist list from ROTA_RULES.md
SPECIALISTS = [
    "Daryl", "Becky", "Billy", "Jinny", "Ginny", "Ben", "Faye", 
    "Claire", "Jake", "Retno", "Jacinta", "Sunny", "Mr Ben", "Nick C"
]

# Specialist subjects that make a FORM teacher free
FREE_SUBJECTS = ['thai', 'music', 'pe', 'p.e.', 'phse']

def clean_name(name):
    if not name: return ""
    n = str(name).strip()
    # Remove title prefixes
    n = re.sub(r'^(mr|mrs|ms|miss|k\.|kun\s|k\s)\s*', '', n, flags=re.IGNORECASE).strip()
    # Remove brackets
    n = n.split('(')[0].strip()
    # Mappings
    nl = n.lower()
    if "jinny" in nl or "ginny" in nl: return "Jinny"
    return n

def init_firestore():
    if not os.path.exists(CREDS_PATH):
        print(f"Error: Credential file {CREDS_PATH} not found.")
        return None
    try:
        if not firebase_admin._apps:
            cred = credentials.Certificate(CREDS_PATH)
            firebase_admin.initialize_app(cred)
        return firestore.client()
    except Exception as e:
        print(f"Firestore Init Error: {e}")
        return None

def target_update_p1():
    db = init_firestore()
    if not db: return

    print(f"Loading Excel: {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    
    # Get all staff from Firestore
    print("Fetching staff list from Firestore...")
    staff_docs = db.collection("staff").get()
    
    update_count = 0
    
    for doc in staff_docs:
        staff_id = doc.id
        staff_data = doc.to_dict()
        original_name = staff_data.get("name", "")
        staff_name = clean_name(original_name)
        
        if not staff_name: continue
        
        # Try to find sheet
        target_sheet = None
        for sname in sheet_names:
            if staff_name.lower() in sname.lower() or sname.lower() in staff_name.lower():
                target_sheet = sname
                break
        
        if not target_sheet:
            # Special case for Claire (ME)
            if staff_name == "Claire" and "ME" in sheet_names:
                target_sheet = "ME"
            else:
                continue

        print(f"Processing P1 for: {original_name} (Sheet: {target_sheet})")
        sheet = wb[target_sheet]
        
        # 1. Find columns
        day_cols = {}
        header_found = False
        for row in sheet.iter_rows(max_row=10, values_only=True):
            for i, cell in enumerate(row):
                if cell:
                    c_val = str(cell).strip().lower()
                    for d in DAYS_LIST:
                        if d.lower() in c_val:
                            day_cols[d] = i
                            header_found = True
            if header_found: break
        
        if not day_cols:
            # Fallback to standard Mon-Fri columns if header not detected
            day_cols = {d: i+1 for i, d in enumerate(DAYS_LIST)}

        # 2. Find P1 row
        p1_row_data = None
        for row in sheet.iter_rows(max_row=60, values_only=True):
            # Check first 5 cols for P1 markers
            for cell in row[:5]:
                if cell:
                    c_val = str(cell).strip().lower()
                    # Regex for "Period 1" or standalone "1"
                    if re.search(r'(^|\b)(period\s*|p\.?\s*|)1(\b|$)', c_val):
                        p1_row_data = row
                        break
            if p1_row_data: break
        
        if p1_row_data:
            batch = db.batch()
            for day, col_idx in day_cols.items():
                if col_idx < len(p1_row_data):
                    activity = str(p1_row_data[col_idx]).strip() if p1_row_data[col_idx] else ""
                    
                    # Logic for is_free
                    is_free = False
                    clean_act = activity.lower()
                    
                    if not clean_act or any(x in clean_act for x in ['free', 'available', 'none', 'nan']):
                        is_free = True
                    elif staff_name not in SPECIALISTS:
                        # Form teacher free if class with specialist
                        if any(sub in clean_act for sub in FREE_SUBJECTS):
                            is_free = True
                    
                    # Update Firestore
                    sched_id = f"{day}_1"
                    sched_ref = db.collection("staff").document(staff_id).collection("schedules").document(sched_id)
                    batch.set(sched_ref, {
                        "day_of_week": day,
                        "period": 1,
                        "activity": activity,
                        "is_free": is_free
                    }, merge=True)
                    update_count += 1
            
            batch.commit()
            print(f"  P1 Sync complete for {original_name}")

    print(f"\n✅ TARGETED UPDATE FINISHED. {update_count} schedule records updated.")

if __name__ == "__main__":
    target_update_p1()
