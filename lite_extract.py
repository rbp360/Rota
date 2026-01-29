import openpyxl
import os

file_path = r"c:\Users\rob_b\Rota\GV cover and staff absence.xlsx"

def lite_extract():
    try:
        print(f"Opening workbook: {file_path}")
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        print(f"Sheets: {wb.sheetnames}")
        
        with open("lite_info.txt", "w") as f:
            f.write(f"Sheets: {wb.sheetnames}\n")
            for name in wb.sheetnames[:5]: # Check first 5 sheets
                f.write(f"\nSheet: {name}\n")
                sheet = wb[name]
                for row in sheet.iter_rows(max_row=5, max_col=10, values_only=True):
                    f.write(f"{row}\n")
        print("Lite extraction complete.")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    lite_extract()
