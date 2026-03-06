import openpyxl

file_path = 'GV cover and staff absence.xlsx'

def find_text_in_excel(text):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    results = []
    for name in wb.sheetnames:
        sheet = wb[name]
        for r_idx, row in enumerate(sheet.iter_rows(max_row=10, max_col=10, values_only=True)):
            for c_idx, val in enumerate(row):
                if val and text.lower() in str(val).lower():
                    results.append(f"Sheet '{name}', Row {r_idx+1}, Col {c_idx+1}: {val}")
                    break
    return results

if __name__ == "__main__":
    print("Searching for Alex...")
    print('\n'.join(find_text_in_excel("Alex")))
    print("\nSearching for Amanda...")
    print('\n'.join(find_text_in_excel("Amanda")))
