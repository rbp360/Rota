import openpyxl
import os

EXCEL_PATH = r"c:\Users\rob_b\Rota\temp_rota.xlsx"
OUTPUT_PATH = r"c:\Users\rob_b\Rota\backend\absence_debug.txt"

def inspect_absence_record():
    with open(OUTPUT_PATH, "w") as f:
        f.write("Starting inspection...\n")
        if not os.path.exists(EXCEL_PATH):
            f.write(f"File not found: {EXCEL_PATH}\n")
            return

        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        sheet_name = 'Absence Record'
        if sheet_name not in wb.sheetnames:
            matches = [s for s in wb.sheetnames if 'absence' in s.lower()]
            if matches:
                sheet_name = matches[0]
                f.write(f"Using sheet: {sheet_name}\n")
            else:
                f.write(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}\n")
                return

        sheet = wb[sheet_name]
        f.write(f"--- Inspecting {sheet_name} ---\n")
        for r_idx, row in enumerate(sheet.iter_rows(max_row=50, values_only=True)):
            f.write(f"Row {r_idx}: {row}\n")
        f.write("Done.\n")

if __name__ == "__main__":
    inspect_absence_record()
