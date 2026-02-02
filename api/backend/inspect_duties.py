
import pandas as pd
import openpyxl

EXCEL_PATH = r"c:\Users\rob_b\Rota\temp_rota.xlsx"

def inspect_duties():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        print(f"Sheet names: {wb.sheetnames}")
        
        for name in ["TB1", "EY"]:
            if name in wb.sheetnames:
                print(f"\n--- Inspecting {name} ---")
                sheet = wb[name]
                # Print first 20 rows to understand layout
                for i, row in enumerate(sheet.iter_rows(max_row=30, values_only=True)):
                    print(f"Row {i+1}: {row}")
            else:
                print(f"\nSheet {name} not found!")


    except Exception as e:
        with open(r"c:\Users\rob_b\Rota\backend\duties_output.txt", "w") as f:
            f.write(f"Error: {e}")

if __name__ == "__main__":
    import sys
    sys.stdout = open(r"c:\Users\rob_b\Rota\backend\duties_output.txt", "w")
    inspect_duties()

