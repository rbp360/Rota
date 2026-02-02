import pandas as pd
import openpyxl
import sys
import os
import re

# Allow running directly as a script
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from backend.database import SessionLocal, Staff, Schedule, Setting, engine, Base
from sqlalchemy.orm import Session
from sqlalchemy import func


EXCEL_PATH = r"c:\Users\rob_b\Rota\temp_rota.xlsx"

def clean_staff_name(name):
    if not name: return ""
    n = str(name).strip()
    
    # GLOBAL IGNORE LIST
    IGNORED = [
        "TBC", "External", "Coach", "Room", "Music Room", "Hall",
        "Gym", "Pitch", "Court", "Pool", "Library", "PRE NURSERY", "PRE NUSERY", "Outside Prov.",
        "**", "gate", "locked", "at", "8.30", "Mr", "1", "Calire", "?"
    ]
    if any(i.lower() in n.lower() for i in IGNORED):
        return None

    # Remove "K.", "Kun ", "K " prefixes
    n = re.sub(r'^(k\.|kun\s|k\s)', '', n, flags=re.IGNORECASE).strip()
    # Remove brackets and content e.g. "Charlotte (Thu)" -> "Charlotte"
    n = n.split('(')[0].strip()
    
    # Common Mappings
    nl = n.lower()
    
    # Fix Typos / Combine Duplicates
    if "jactina" in nl: return "Jacinta"
    if "nokkeaw" in nl: return "Nokkaew"
    if "nick" in nl and ("c" in nl or nl == "nick"): return "Nick C" # Matches "Nick C", "Nick. C", "Nick"
    
    if nl == "darryl": return "Daryl"
    if nl == "ginny": return "Jinny"
    if nl == "jinny": return "Jinny" 
    if nl == "janel": return "Janel"
    
    return n

def normalize_data():
    db = SessionLocal()
    try:
        # Clear existing data
        db.query(Schedule).delete()
        db.query(Staff).delete()
        db.commit()

        days_list = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        
        print(f"Loading workbook: {EXCEL_PATH}...")
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        sheet_names = wb.sheetnames

        teacher_profiles = {
            "Daryl": "Music teacher",
            "Jake": "Assistant that works throughout school",
            "Becky": "Drama teacher who teaches whole school",
            "Billy": "PE teacher who teaches whole school",
            "Retno": "Pre-nursery teacher",
            "Jacinta": "Nursery teacher",
            "Faye": "Predominantly used for covering",
            "Ginny": "Assistant who works with different classes and students",
            "Claire": "Head (used for cover), tab is 'ME'",
            "Ben": "Forest school teacher who teaches whole classes",
            "Baitoey": "Teaching Assistant (Duty Only)",
            "Nop": "Teaching Assistant (Duty Only)",
            "Tum": "Teaching Assistant (Duty Only)",
            "Nick C": "Qualified Teacher (Duty + Periods)",
            "Kat": "Teaching Assistant (Duty Only)",
            "Mr Ben": "Qualified Teacher (Periods + Duties)",
            "Janel": "Teaching Assistant"
        }

        # Staff with full period cover capability (Qualified Teachers)
        duty_only_staff = ["Baitoey", "Nop", "Tum", "Kat", "Janel"]
        
        # Identify specialists (non-form teachers)
        specialists_list = [
            "Daryl", "Becky", "Billy", "Jinny", "Ginny", "Ben", "Faye", 
            "Claire", "Jake", "Retno", "Jacinta", "Sunny", "Mr Ben", "Nick C"
        ]

        for name in sheet_names:
            name_lower = name.lower().replace(" ", "")

            ignore_list = [
                'instructions', 'summary', 'record', 'absencerecord', 
                'sheet3', 'sheet1', 'cca', 'y56eal'
            ]
            if name_lower in ignore_list or name_lower.startswith('sheet') or name_lower in ['tb1', 'ey']:
                continue
            
            staff_name = clean_staff_name(name)
            if not staff_name: 
                print(f"Skipping ignored sheet: {name}")
                continue

            if name == "ME": staff_name = "Claire"
            
            print(f"--- Normalizing: {staff_name} ---")
            sheet = wb[name]
            
            is_spec = (staff_name in specialists_list)
            can_cover = staff_name not in duty_only_staff

            staff = db.query(Staff).filter(func.lower(Staff.name) == staff_name.lower()).first()
            if not staff:
                staff = Staff(
                    name=staff_name,
                    role="Teacher" if "Assistant" not in teacher_profiles.get(staff_name, "") and staff_name not in duty_only_staff else "TA",
                    profile=teacher_profiles.get(staff_name, ""),
                    is_priority=(staff_name == "Claire"),
                    is_specialist=is_spec,
                    is_active=True,
                    can_cover_periods=can_cover
                )
                db.add(staff)
                db.flush()


            # Find day columns - scan first 20 rows
            day_cols = {}
            header_row_idx = 0
            for r_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
                found_in_row = False
                for c_idx, val in enumerate(row):
                    if not val: continue
                    val_str = str(val).strip().lower()
                    for d in days_list:
                        if d.lower() == val_str: # Exact match
                            day_cols[d] = c_idx + 1
                            found_in_row = True
                        elif d.lower() in val_str and len(val_str) < 15: # Partial match (e.g. "Monday 25th")
                            day_cols[d] = c_idx + 1
                            found_in_row = True
                if found_in_row and len(day_cols) >= 3:
                    header_row_idx = r_idx + 1
                    break
            
            if not day_cols:
                day_cols = {d: i+2 for i, d in enumerate(days_list)}
                header_row_idx = 1
            
            print(f"  Day Columns: {day_cols} (Header Row: {header_row_idx})")

            # Parse periods 1-8
            p_rows_map = {}
            for r_idx, row in enumerate(sheet.iter_rows(max_row=60, values_only=True)):
                # Check first 5 columns for period markers
                markers = [str(v).strip().lower() for v in row[:5] if v is not None]
                for p_num in range(1, 9):
                    if p_num in p_rows_map: continue
                    # Common markers: "1", "P1", "Period 1", "P.1"
                    possible = [str(p_num), f"p{p_num}", f"period {p_num}", f"p.{p_num}", f"per {p_num}"]
                    if any(m in markers for m in possible):
                        p_rows_map[p_num] = row
                        break
            
            for p_num in range(1, 9):
                row_data = p_rows_map.get(p_num)
                if row_data is None:
                    # Fallback
                    rows_cached = list(sheet.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 30, values_only=True))
                    row_data = rows_cached[p_num-1] if (p_num-1) < len(rows_cached) else [None]*50
                    print(f"    P{p_num} using fallback row")
                else:
                    print(f"    P{p_num} found marker row")

                for day, col in day_cols.items():
                    raw_val = row_data[col-1] if col <= len(row_data) else None
                    val = str(raw_val).strip() if raw_val is not None else ""
                    
                    is_available = False
                    clean_val = val.lower().replace(" ", "")
                    free_keywords = ['none', 'nan', 'free', 'available', '0', '0.0', '']
                    
                    # New Rules: Thai, Music, PE, PHSE free form teachers
                    specialist_subjects = ['thai', 'music', 'pe', 'p.e.', 'phse']
                    is_specialist_lesson = any(sub in val.lower() for sub in specialist_subjects)
                    
                    if not clean_val or clean_val in free_keywords:
                        is_available = True
                    elif not is_spec and is_specialist_lesson:
                        # If NOT a specialist teacher, but doing a specialist subject, they are free!
                        is_available = True
                    
                    db.add(Schedule(
                        staff_id=staff.id,
                        day_of_week=day,
                        period=p_num,
                        activity=val,
                        is_free=is_available
                    ))
        
        # Ensure all profile staff exist (for those without sheets)
        for p_name, p_desc in teacher_profiles.items():
             s_name = clean_staff_name(p_name)
             if not s_name: continue
             
             existing = db.query(Staff).filter(func.lower(Staff.name) == s_name.lower()).first()
             if not existing:
                 can_cover = s_name not in duty_only_staff
                 staff = Staff(
                    name=s_name,
                    role="TA" if "Assistant" in p_desc or s_name in duty_only_staff else "Teacher",
                    profile=p_desc,
                    is_priority=(s_name == "Claire"),
                    is_specialist=(s_name in specialists_list),
                    is_active=True,
                    can_cover_periods=can_cover
                 )
                 db.add(staff)
                 print(f"Added missing profile staff: {s_name}")
        
        db.commit()

        # Process Duty Sheets
        for duty_sheet in ['TB1', 'EY']:
            if duty_sheet not in wb.sheetnames:
                continue
            
            print(f"--- Processing Duties: {duty_sheet} ---")
            sheet = wb[duty_sheet]
            
            # Find Day Columns
            # Expected Structure: [Desc] [Time] [Type] [Duration] [Mon] [Tue] [Wed] [Thu] [Fri]
            # We look for the header row containing "Monday"
            header_row_idx = None
            day_cols = {}
            for r_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
                row_str = [str(c).lower() for c in row if c]
                if "monday" in row_str:
                    header_row_idx = r_idx + 1
                    for c_idx, val in enumerate(row):
                        if not val: continue
                        val_str = str(val).strip().lower()
                        for d in days_list:
                            if d.lower() in val_str:
                                day_cols[d] = c_idx + 1
                    break
            
            # Fallback if no header found (based on user snippet inference)
            if not day_cols:
                # Assuming Cols E, F, G, H, I (indices 5,6,7,8,9) -> Python 0-indexed: 4,5,6,7,8
                # But openpyxl is 1-indexed for col numbers? No, iter_rows vals are tuple.
                # Let's assume standard excel layout: A=1, B=2... E=5.
                day_cols = {
                    'Monday': 5, 'Tuesday': 6, 'Wednesday': 7, 'Thursday': 8, 'Friday': 9
                }
                header_row_idx = 1 # Guess
                print("  Using fallback column indices for Duties (E-I)")

            print(f"  Duty Day Columns: {day_cols}")

            # Iterate rows
            for r_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx+1, values_only=True)):
                if not row or all(c is None for c in row): continue
                
                # Determine Duty Period based on Time (Col B / idx 1) or Type (Col C / idx 2)
                # Col A=0, B=1, C=2
                time_val = str(row[1]).lower() if len(row) > 1 and row[1] else ""
                type_val = str(row[2]).lower() if len(row) > 2 and row[2] else ""
                desc_val = str(row[0]).lower() if len(row) > 0 and row[0] else ""
                
                combined_marker = f"{time_val} {type_val} {desc_val}"
                
                period_num = 9 # Default Lunch
                duty_name = "Lunch Duty"
                
                if "8." in combined_marker or "08" in combined_marker or "before" in combined_marker:
                    period_num = 0
                    duty_name = "Before School Duty"
                elif "10." in combined_marker or "break" in combined_marker:
                    period_num = 11 # Break
                    duty_name = "Break Duty"
                elif "12." in combined_marker or "1." in combined_marker or "lunch" in combined_marker:
                    period_num = 9 # Lunch
                    duty_name = "Lunch Duty"
                elif "15." in combined_marker or "3." in combined_marker or "after" in combined_marker:
                    period_num = 10
                    duty_name = "After School Duty"

                # Extract staff from day columns
                for day, col_idx in day_cols.items():
                    if col_idx > len(row): continue
                    
                    # col_idx is 1-based from earlier logic? 
                    # If I manually set 5, it means 5th column. row tuple is 0-indexed.
                    # so row[col_idx-1]
                    cell_val = row[col_idx-1]
                    if not cell_val: continue
                    val_str = str(cell_val).strip()
                    if val_str.lower() in ['none', 'nan', '']: continue
                    
                    # Logic for "Claire (Tue) + Faye (Mon)" type cells
                    # If the cell contains brackets with day names, we filter.
                    # Otherwise we assume it applies to THIS day column (if in a grid)
                    # OR if it's a merged cell spanning all, we check days?
                    # The user snippet showed "Claire (Tuesday...)" inside a cell.
                    
                    target_staff_names = []
                    
                    # Split by newlines, "+", "&"
                    import re
                    # specific cleanup for "Claire (Tuesday...)" pattern
                    # If specific days are mentioned in brackets, only assign if matches current 'day'
                    
                    if "(" in val_str and any(d[:3].lower() in val_str.lower() for d in days_list):
                        # Complex cell
                        # Check if CURRENT day is mentioned
                        if day.lower() in val_str.lower() or day[:3].lower() in val_str.lower():
                            # Who is associated with this day?
                            # Regex: Name \((...)\)
                            # This is hard to parse perfectly without more robust NLP, but let's try strict tokenizing
                            # "Claire (Tuesday, Wednesday)" -> if today is Tuesday, add Claire.
                            # We can just look for the name preceding the bracket containing the day?
                            # Or simpler: Just look for names in the string.
                            # If a name is found, blindly assign it? NO, wrong day.
                            
                            # Heuristic: Split by "+" or newline.
                            parts = re.split(r'[+\n]', val_str)
                            for part in parts:
                                part = part.strip()
                                if not part: continue
                                
                                part_lower = part.lower()
                                # Check if this part has ANY day mentioned
                                # (Use full day names or 3-letter abbr)
                                days_mentioned_in_part = False
                                for d in days_list:
                                    if d.lower() in part_lower or d[:3].lower() in part_lower:
                                        days_mentioned_in_part = True
                                        break
                                
                                # Condition to assign:
                                # 1. The current day is explicitly mentioned in this part
                                # 2. OR No days are mentioned in this part (implies generic assignment for this cell's column)
                                current_day_match = (day.lower() in part_lower) or (day[:3].lower() in part_lower)
                                
                                if current_day_match or not days_mentioned_in_part:
                                    # Extract name from this part (remove brackets and content)
                                    name_part = part.split('(')[0].strip()
                                    if name_part:
                                        target_staff_names.append(name_part)

                    else:
                        # Simple cell (just names)
                        names = re.split(r'[+\n&,\s]', val_str) # Added space to splitters
                        for n in names:
                            n = n.strip()
                            if n: target_staff_names.append(n)
                    
                    # Register Staff and Assignments
                    for raw_name in target_staff_names:
                        map_name = clean_staff_name(raw_name)
                        if not map_name: continue
                        
                        # Find or Create
                        staff_obj = db.query(Staff).filter(func.lower(Staff.name) == map_name.lower()).first()
                        
                        if not staff_obj:
                             # New staff found in Duty
                             is_qts = map_name in ["Mr Ben", "Nick C"]
                             staff_obj = Staff(
                                 name=map_name,
                                 role="Teacher" if is_qts else "Duties Only",
                                 profile="Added from Duty Rota",
                                 is_active=True,
                                 is_specialist=False,
                                 can_cover_periods=is_qts # False for most duties
                             )
                             db.add(staff_obj)
                             db.flush() # get ID
                             print(f"    New Staff from Duty: {map_name} (Teacher={is_qts})")
                        
                        # Add Schedule
                        db.add(Schedule(
                            staff_id=staff_obj.id,
                            day_of_week=day,
                            period=period_num,
                            activity=f"{duty_name}: {desc_val}",
                            is_free=False
                        ))
        
        db.commit()

        # Process CCA Sheet
        if "CCA" in wb.sheetnames:
            print("--- Processing CCA (Smart Pairing) ---")
            sheet = wb["CCA"]
            
            # 1. Identify Day Columns and Header Row
            day_cols_map = {} # day -> { 'act_col': idx, 'staff_col': idx }
            header_row_idx = None
            
            # Scan top 20 rows for "Monday", "Tuesday" etc.
            for r_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
                row_vals = [str(c).lower() if c else "" for c in row]
                found_days = []
                for d in days_list:
                    if d.lower() in row_vals:
                        found_days.append(d)
                
                if found_days:
                    header_row_idx = r_idx + 1
                    # Map Day name to the pair of columns
                    # User: "Club names appear in the columns first and the teachers second"
                    # We look for the day name. If it's found at idx, 
                    # then idx is usually the Activity and idx+1 is Staff.
                    for d in found_days:
                        c_idx = row_vals.index(d.lower())
                        day_cols_map[d] = {
                            'act_col': c_idx,
                            'staff_col': c_idx + 1
                        }
                    break
            
            if not day_cols_map:
                 print("  Could not find day headers in CCA tab.")
            else:
                 print(f"  Detected CCA Mapping: {day_cols_map}")
                 # Iterate CCA data rows
                 for row_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx+1, values_only=True)):
                     if not any(row): continue
                     
                     for day, col_cfg in day_cols_map.items():
                         a_idx = col_cfg['act_col']
                         s_idx = col_cfg['staff_col']
                         
                         if s_idx >= len(row): continue
                         
                         raw_cca = row[a_idx]
                         raw_staff = row[s_idx]
                         
                         if not raw_staff or str(raw_staff).lower() in ['none', 'nan', '']: 
                             continue
                         
                         # Activity name comes from the first column of the pair
                         cca_name = str(raw_cca).strip() if raw_cca else "CCA"
                         staff_names_str = str(raw_staff).strip()
                         
                         # Split multiple staff in a cell
                         raw_names = re.split(r'[+=\n&]', staff_names_str)
                         
                         for raw_n in raw_names:
                             clean_name = clean_staff_name(raw_n)
                             if not clean_name: continue
                             
                             # FILTER: Ignore Secondary
                             if "(sec)" in raw_n.lower(): continue

                             # FILTER: Explicit Ignore List (Now handled globally in clean_staff_name)
                             # IGNORED_STAFF block removed
                             
                             # Find or Create
                             staff_obj = db.query(Staff).filter(func.lower(Staff.name) == clean_name.lower()).first()
                             
                             if not staff_obj:
                                 # This is a new staff member found only in CCA
                                 staff_obj = Staff(
                                     name=clean_name,
                                     role="TA (from CCA)",
                                     profile="Added from CCA Rota",
                                     is_active=True,
                                     is_specialist=False,
                                     can_cover_periods=False
                                 )
                                 db.add(staff_obj)
                                 db.flush()
                                 print(f"    New Staff from CCA: {clean_name}")
                             
                             # Add the CCA to their schedule
                             db.add(Schedule(
                                 staff_id=staff_obj.id,
                                 day_of_week=day,
                                 period=13, # CCA Period
                                 activity=f"CCA: {cca_name}",
                                 is_free=False
                             ))
                             print(f"    Assigned {clean_name} -> {cca_name} ({day})")

        # Final Deduplication Safety Net
        print("--- Final Deduplication Check ---")
        all_staff = db.query(Staff).all()
        seen = {} # canonical_name -> staff_obj
        for s in all_staff:
            canon = clean_staff_name(s.name)
            if not canon: 
                print(f"Deleting ignored staff: {s.name}")
                db.query(Schedule).filter(Schedule.staff_id == s.id).delete()
                db.query(Absence).filter(Absence.staff_id == s.id).delete()
                db.query(Cover).filter(Cover.covering_staff_id == s.id).delete()
                db.delete(s)
                continue
            
            if canon in seen:
                # Duplicate! Merge schedules, absences, and covers
                primary = seen[canon]
                print(f"MERGING DUPLICATE: {s.name} into {primary.name}")
                
                # Update Schedules
                db.query(Schedule).filter(Schedule.staff_id == s.id).update({Schedule.staff_id: primary.id})
                
                # Update Absences
                db.query(Absence).filter(Absence.staff_id == s.id).update({Absence.staff_id: primary.id})
                
                # Update Covers
                db.query(Cover).filter(Cover.covering_staff_id == s.id).update({Cover.covering_staff_id: primary.id})
                
                db.delete(s)
            else:
                seen[canon] = s
                # Ensure name is canonical
                if s.name != canon:
                    s.name = canon

        db.commit()
        print("Normalization complete.")
        
    except Exception as e:
        print(f"Error during normalization: {e}")
        db.rollback()
    finally:
        db.close()

if __name__ == "__main__":
    # Ensure tables exist
    Base.metadata.create_all(bind=engine)
    normalize_data()
