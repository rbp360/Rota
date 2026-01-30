import openpyxl
import os
import re
import sqlite3
from datetime import datetime

EXCEL_PATH = r"c:\Users\rob_b\Rota\temp_rota.xlsx"
DB_PATH = r"c:\Users\rob_b\Rota\rota.db"
LOG_PATH = r"c:\Users\rob_b\Rota\backend\legacy_direct_log.txt"

def clean_staff_name(name):
    if not name: return None
    n = str(name).strip().split('(')[0].strip()
    n = re.sub(r'(\d+\.\d+|pm|am|late|from|0\.5|half|\.|\?)', '', n, flags=re.IGNORECASE).strip()
    if not n or len(n) < 2: return None
    
    nl = n.lower()
    if "jactina" in nl: return "Jacinta"
    if "nokkeaw" in nl: return "Nokkaew"
    if "nick" in nl and ("c" in nl or nl == "nick"): return "Nick C"
    if "soe" in nl and len(nl) < 6: return "Soe"
    if nl == "darryl": return "Daryl"
    return n

def parse_date(date_str):
    match = re.search(r'(\d+)/(\d+)', date_str)
    if not match: return None
    day, month = int(match.group(1)), int(match.group(2))
    year = 2025 if month >= 8 else 2026
    try: return f"{year}-{month:02d}-{day:02d}"
    except: return None

def run_direct():
    with open(LOG_PATH, "w") as log:
        log.write("Starting direct legacy normalization...\n")
        try:
            conn = sqlite3.connect(DB_PATH)
            cursor = conn.cursor()
            
            # Get staff mapping
            cursor.execute("SELECT id, name FROM staff")
            staff_map = {row[1].lower(): row[0] for row in cursor.fetchall()}
            log.write(f"Loaded {len(staff_map)} staff from DB.\n")
            
            wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
            sheet = wb['Absence Record'] if 'Absence Record' in wb.sheetnames else wb[wb.sheetnames[0]] # fallback
            
            days_pattern = re.compile(r'(Monday|Tuesday|Wednesday|Thursday|Friday)', re.IGNORECASE)
            
            count = 0
            for row in sheet.iter_rows(values_only=True):
                for c_idx, cell in enumerate(row):
                    if not cell: continue
                    val = str(cell).strip()
                    if days_pattern.search(val) and '/' in val:
                        target_date = parse_date(val)
                        if not target_date: continue
                        
                        staff_cell_val = ""
                        for offset in range(1, 4):
                            if c_idx + offset < len(row):
                                next_cell = row[c_idx + offset]
                                if next_cell and not days_pattern.search(str(next_cell)) and '/' not in str(next_cell):
                                    staff_cell_val = str(next_cell)
                                    break
                        
                        if not staff_cell_val: continue
                        
                        is_pm = any(x in staff_cell_val.lower() for x in ["pm", "from 1", "from 12"])
                        is_half = any(x in staff_cell_val.lower() for x in ["am", "pm", "late", "from", "0.5", "half"])
                        start_p, end_p = (1, 8)
                        if is_half: start_p, end_p = (5, 8) if is_pm else (1, 4)
                        
                        s_clean = re.sub(r'(&|and|\+)', ',', staff_cell_val, flags=re.IGNORECASE)
                        names = [n.strip() for n in re.split(r'[,\n]', s_clean)]
                        
                        for raw_n in names:
                            cn = clean_staff_name(raw_n)
                            if not cn: continue
                            
                            sid = staff_map.get(cn.lower())
                            if not sid:
                                # Loose match
                                match = [s_id for s_name, s_id in staff_map.items() if s_name in cn.lower() or cn.lower() in s_name]
                                if match: sid = match[0]
                                else: continue
                            
                            cursor.execute("SELECT id FROM absences WHERE staff_id=? AND date=?", (sid, target_date))
                            if not cursor.fetchone():
                                cursor.execute("INSERT INTO absences (staff_id, date, start_period, end_period, reason) VALUES (?, ?, ?, ?, ?)",
                                             (sid, target_date, start_p, end_p, f"Legacy: {staff_cell_val[:50]}"))
                                count += 1
                                log.write(f"ADDED: {cn} on {target_date}\n")
            
            conn.commit()
            log.write(f"Done. Added {count} entries.\n")
            conn.close()
        except Exception as e:
            log.write(f"ERROR: {e}\n")

if __name__ == "__main__":
    run_direct()
