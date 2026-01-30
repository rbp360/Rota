import openpyxl
import os
import re
import pandas as pd
from datetime import datetime
from .database import SessionLocal, Staff, Absence, engine, Base
from sqlalchemy.orm import Session
from sqlalchemy import func

EXCEL_PATH = r"c:\Users\rob_b\Rota\temp_rota.xlsx"

def clean_staff_name(name):
    if not name: return None
    n = str(name).strip().split('(')[0].strip()
    
    # Global ignore list from normalize.py
    IGNORED = [
        "TBC", "External", "Coach", "Room", "Music Room", "Hall",
        "Gym", "Pitch", "Court", "Pool", "Library", "PRE NURSERY", "PRE NUSERY", "Outside Prov.",
        "**", "gate", "locked", "at", "8.30", "Mr", "1", "Calire", "?", "pd", "consulate", "cons", "pl", "dl"
    ]
    
    # Remove common suffixes/patterns
    n = re.sub(r'(\d+\.\d+|pm|am|late|from|0\.5|half|\.|\?)', '', n, flags=re.IGNORECASE).strip()
    
    if not n or len(n) < 2: return None
    
    if any(i.lower() in n.lower() for i in IGNORED if len(i) > 2):
        return None

    # Common Mappings
    nl = n.lower()
    if "jactina" in nl: return "Jacinta"
    if "nokkeaw" in nl: return "Nokkaew"
    if "nick" in nl and ("c" in nl or nl == "nick"): return "Nick C"
    if "soe" in nl and len(nl) < 6: return "Soe"
    if nl == "darryl": return "Daryl"
    if nl == "ginny": return "Jinny"
    if nl == "anniina": return "Anniina"
    
    return n

def parse_date(date_str):
    match = re.search(r'(\d+)/(\d+)', date_str)
    if not match: return None
    day, month = int(match.group(1)), int(match.group(2))
    year = 2025 if month >= 8 else 2026
    try: return datetime(year, month, day).date()
    except: return None

def normalize_legacy_absences():
    db = SessionLocal()
    try:
        if not os.path.exists(EXCEL_PATH): return

        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        sheet_name = 'Absence Record'
        if sheet_name not in wb.sheetnames:
            matches = [s for s in wb.sheetnames if 'absence' in s.lower()]
            if not matches: return
            sheet_name = matches[0]
        
        sheet = wb[sheet_name]
        all_staff_names = {s.name.lower(): s.id for s in db.query(Staff).all()}
        days_pattern = re.compile(r'(Monday|Tuesday|Wednesday|Thursday|Friday)', re.IGNORECASE)
        
        for row in sheet.iter_rows(values_only=True):
            for c_idx, cell in enumerate(row):
                if not cell: continue
                val = str(cell).strip()
                
                if days_pattern.search(val) and '/' in val:
                    target_date = parse_date(val)
                    if not target_date: continue
                    
                    staff_cell_val = ""
                    for offset in range(1, 4):
                        if c_idx + offset < len(row):
                            next_cell = row[c_idx + offset]
                            if next_cell and not days_pattern.search(str(next_cell)) and '/' not in str(next_cell):
                                staff_cell_val = str(next_cell)
                                break
                    
                    if not staff_cell_val: continue
                    
                    s_orig = staff_cell_val
                    is_pm = any(x in s_orig.lower() for x in ["pm", "from 1", "from 12"])
                    is_half = any(x in s_orig.lower() for x in ["am", "pm", "late", "from", "0.5", "half", "(pd)"])
                    
                    start_p, end_p = (1, 8)
                    if is_half: start_p, end_p = (5, 8) if is_pm else (1, 4)
                    
                    s_clean = re.sub(r'(&|and|\+)', ',', s_orig, flags=re.IGNORECASE)
                    potentials = [p.strip() for p in re.split(r'[,\n]', s_clean)]
                    
                    final_names = []
                    for p in potentials:
                        if not p: continue
                        words = p.split()
                        current_name = ""
                        for w in words:
                            if w.lower() in ["pm", "am", "late", "from", "consulate", "0.5"]: continue
                            test_name = (current_name + " " + w).strip()
                            clean_w = clean_staff_name(w)
                            if clean_w and clean_w.lower() in all_staff_names and current_name:
                                final_names.append(current_name)
                                current_name = w
                            else: current_name = test_name
                        if current_name: final_names.append(current_name)

                    processed_names = set()
                    for raw_n in final_names:
                        cn = clean_staff_name(raw_n)
                        if cn: processed_names.add(cn)
                    
                    for name in processed_names:
                        staff_id = all_staff_names.get(name.lower())
                        if not staff_id:
                            match = [sid for sn, sid in all_staff_names.items() if sn in name.lower() or name.lower() in sn]
                            if match: staff_id = match[0]
                            else: continue
                        
                        exists = db.query(Absence).filter(Absence.staff_id == staff_id, Absence.date == target_date).first()
                        
                        if not exists:
                            db.add(Absence(staff_id=staff_id, date=target_date, start_period=start_p, end_period=end_p, reason=f"Legacy: {s_orig[:50]}"))
        db.commit()
    except Exception as e:
        print(f"Error: {e}")
        db.rollback()
    finally:
        db.close()

if __name__ == "__main__":
    normalize_legacy_absences()
