import pandas as pd
import openpyxl
import sys
import os

# Allow running directly as a script
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from backend.database import SessionLocal, Staff, Schedule, Setting, engine, Base
from sqlalchemy.orm import Session
from sqlalchemy import func


EXCEL_PATH = r"c:\Users\rob_b\Rota\temp_rota.xlsx"

def normalize_data():
    db = SessionLocal()
    try:
        # Clear existing data
        db.query(Schedule).delete()
        db.query(Staff).delete()
        db.commit()

        days_list = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        
        print(f"Loading workbook: {EXCEL_PATH}...")
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        sheet_names = wb.sheetnames

        teacher_profiles = {
            "Daryl": "Music teacher",
            "Jake": "Assistant that works throughout school",
            "Becky": "Drama teacher who teaches whole school",
            "Billy": "PE teacher who teaches whole school",
            "Retno": "Pre-nursery teacher",
            "Jacinta": "Nursery teacher",
            "Faye": "Predominantly used for covering",
            "Ginny": "Assistant who works with different classes and students",
            "Claire": "Head (used for cover), tab is 'ME'",
            "Ben": "Forest school teacher who teaches whole classes",
            "Baitoey": "Teaching Assistant (Duty Only)",
            "Nop": "Teaching Assistant (Duty Only)",
            "Tum": "Teaching Assistant (Duty Only)",
            "Nick.C": "Qualified Teacher (Duty + Periods)",
            "Kat": "Teaching Assistant (Duty Only)",
            "Mr Ben": "Qualified Teacher (Periods + Duties)",
            "Janel": "Teaching Assistant"
        }

        # Staff with full period cover capability (Qualified Teachers)
        # Assuming most from sheet_names are teachers, but we need to restrict new ones.
        # Explicitly list who CANNOT do periods (TAs)
        duty_only_staff = ["Baitoey", "Nop", "Tum", "Kat", "Janel"]

        # Exceptions and Aliases
        # Nick -> Nick.C handled by checking normalized name? 
        # User says "Nick (Nick.C)" - likely means "Nick" in sheet maps to "Nick.C"
        
        # Identify specialists (non-form teachers)
        specialists_list = [
            "Daryl", "Becky", "Billy", "Jinny", "Ginny", "Ben", "Faye", 
            "Claire", "Jake", "Retno", "Jacinta", "Sunny", "Mr Ben", "Nick.C"
        ]

        for name in sheet_names:
            name_lower = name.lower().replace(" ", "")

            ignore_list = [
                'instructions', 'summary', 'record', 'absencerecord', 
                'sheet3', 'sheet1', 'cca', 'y56eal'
            ]
            # remove 'tb1', 'ey' from ignore list if we want to process them separately, 
            # but here we loop for STAFF sheets.
            if name_lower in ignore_list or name_lower.startswith('sheet') or name_lower in ['tb1', 'ey']:
                continue
            
            staff_name = name
            if name == "ME": staff_name = "Claire"
            if name.lower() == "pre nursery": staff_name = "Retno"
            
            # Name Normalization
            if staff_name.lower() == "nick": staff_name = "Nick.C"
            if staff_name.lower() == "janel": staff_name = "Janel" # Ensure case


            print(f"--- Normalizing: {staff_name} ---")
            sheet = wb[name]
            
            # Explicitly force is_specialist to match the list exactly, else False (0)
            is_spec = (staff_name in specialists_list)
            
            # Check if they can cover periods
            can_cover = staff_name not in duty_only_staff

            
            # Handle Mr Ben vs Ben
            # If sheet name is "Ben", it is Ben (Forest School).
            # If user refers to "Mr Ben", we need to ensure he exists.
            
            staff = db.query(Staff).filter(Staff.name == staff_name).first()
            if not staff:
                staff = Staff(
                    name=staff_name,
                    role="Teacher" if "Assistant" not in teacher_profiles.get(staff_name, "") and staff_name not in duty_only_staff else "TA",
                    profile=teacher_profiles.get(staff_name, ""),
                    is_priority=(staff_name == "Claire"),
                    is_specialist=is_spec,
                    is_active=True,
                    can_cover_periods=can_cover
                )
                db.add(staff)
                db.flush()


            # Find day columns - scan first 20 rows
            day_cols = {}
            header_row_idx = 0
            for r_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
                found_in_row = False
                for c_idx, val in enumerate(row):
                    if not val: continue
                    val_str = str(val).strip().lower()
                    for d in days_list:
                        if d.lower() == val_str: # Exact match
                            day_cols[d] = c_idx + 1
                            found_in_row = True
                        elif d.lower() in val_str and len(val_str) < 15: # Partial match (e.g. "Monday 25th")
                            day_cols[d] = c_idx + 1
                            found_in_row = True
                if found_in_row and len(day_cols) >= 3:
                    header_row_idx = r_idx + 1
                    break
            
            if not day_cols:
                day_cols = {d: i+2 for i, d in enumerate(days_list)}
                header_row_idx = 1
            
            print(f"  Day Columns: {day_cols} (Header Row: {header_row_idx})")

            # Parse periods 1-8
            p_rows_map = {}
            for r_idx, row in enumerate(sheet.iter_rows(max_row=60, values_only=True)):
                # Check first 5 columns for period markers
                markers = [str(v).strip().lower() for v in row[:5] if v is not None]
                for p_num in range(1, 9):
                    if p_num in p_rows_map: continue
                    # Common markers: "1", "P1", "Period 1", "P.1"
                    possible = [str(p_num), f"p{p_num}", f"period {p_num}", f"p.{p_num}", f"per {p_num}"]
                    if any(m in markers for m in possible):
                        p_rows_map[p_num] = row
                        break
            
            for p_num in range(1, 9):
                row_data = p_rows_map.get(p_num)
                if row_data is None:
                    # Fallback
                    rows_cached = list(sheet.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 30, values_only=True))
                    row_data = rows_cached[p_num-1] if (p_num-1) < len(rows_cached) else [None]*50
                    print(f"    P{p_num} using fallback row")
                else:
                    print(f"    P{p_num} found marker row")

                for day, col in day_cols.items():
                    raw_val = row_data[col-1] if col <= len(row_data) else None
                    val = str(raw_val).strip() if raw_val is not None else ""
                    
                    is_available = False
                    clean_val = val.lower().replace(" ", "")
                    free_keywords = ['none', 'nan', 'free', 'available', '0', '0.0', '']
                    
                    # New Rules: Thai, Music, PE, PHSE free form teachers
                    specialist_subjects = ['thai', 'music', 'pe', 'p.e.', 'phse']
                    is_specialist_lesson = any(sub in val.lower() for sub in specialist_subjects)
                    
                    if not clean_val or clean_val in free_keywords:
                        is_available = True
                    elif not is_spec and is_specialist_lesson:
                        # If NOT a specialist teacher, but doing a specialist subject, they are free!
                        is_available = True
                    
                    db.add(Schedule(
                        staff_id=staff.id,
                        day_of_week=day,
                        period=p_num,
                        activity=val,
                        is_free=is_available
                    ))
        
        # Ensure all profile staff exist (for those without sheets)
        for p_name, p_desc in teacher_profiles.items():
             # Normalize name
             s_name = p_name
             if s_name.lower() == "nick": s_name = "Nick.C"
             
             existing = db.query(Staff).filter(Staff.name == s_name).first()
             if not existing:
                 can_cover = s_name not in duty_only_staff
                 staff = Staff(
                    name=s_name,
                    role="TA" if "Assistant" in p_desc or s_name in duty_only_staff else "Teacher",
                    profile=p_desc,
                    is_priority=(s_name == "Claire"),
                    is_specialist=(s_name in specialists_list), # specialists_list needs to be accessible
                    is_active=True,
                    can_cover_periods=can_cover
                 )
                 db.add(staff)
                 print(f"Added missing staff: {s_name}")
        
        db.commit()

        # Process Duty Sheets
        for duty_sheet in ['TB1', 'EY']:
            if duty_sheet not in wb.sheetnames:
                continue
            
            print(f"--- Processing Duties: {duty_sheet} ---")
            sheet = wb[duty_sheet]
            
            # Find Day Columns
            # Expected Structure: [Desc] [Time] [Type] [Duration] [Mon] [Tue] [Wed] [Thu] [Fri]
            # We look for the header row containing "Monday"
            header_row_idx = None
            day_cols = {}
            for r_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
                row_str = [str(c).lower() for c in row if c]
                if "monday" in row_str:
                    header_row_idx = r_idx + 1
                    for c_idx, val in enumerate(row):
                        if not val: continue
                        val_str = str(val).strip().lower()
                        for d in days_list:
                            if d.lower() in val_str:
                                day_cols[d] = c_idx + 1
                    break
            
            # Fallback if no header found (based on user snippet inference)
            if not day_cols:
                # Assuming Cols E, F, G, H, I (indices 5,6,7,8,9) -> Python 0-indexed: 4,5,6,7,8
                # But openpyxl is 1-indexed for col numbers? No, iter_rows vals are tuple.
                # Let's assume standard excel layout: A=1, B=2... E=5.
                day_cols = {
                    'Monday': 5, 'Tuesday': 6, 'Wednesday': 7, 'Thursday': 8, 'Friday': 9
                }
                header_row_idx = 1 # Guess
                print("  Using fallback column indices for Duties (E-I)")

            print(f"  Duty Day Columns: {day_cols}")

            # Iterate rows
            for r_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx+1, values_only=True)):
                if not row or all(c is None for c in row): continue
                
                # Determine Duty Period based on Time (Col B / idx 1) or Type (Col C / idx 2)
                # Col A=0, B=1, C=2
                time_val = str(row[1]).lower() if len(row) > 1 and row[1] else ""
                type_val = str(row[2]).lower() if len(row) > 2 and row[2] else ""
                desc_val = str(row[0]).lower() if len(row) > 0 and row[0] else ""
                
                combined_marker = f"{time_val} {type_val} {desc_val}"
                
                period_num = 9 # Default Lunch
                duty_name = "Lunch Duty"
                
                if "8." in combined_marker or "08" in combined_marker or "before" in combined_marker:
                    period_num = 0
                    duty_name = "Before School Duty"
                elif "10." in combined_marker or "break" in combined_marker:
                    period_num = 11 # Break
                    duty_name = "Break Duty"
                elif "12." in combined_marker or "1." in combined_marker or "lunch" in combined_marker:
                    period_num = 9 # Lunch
                    duty_name = "Lunch Duty"
                elif "15." in combined_marker or "3." in combined_marker or "after" in combined_marker:
                    period_num = 10
                    duty_name = "After School Duty"

                # Extract staff from day columns
                for day, col_idx in day_cols.items():
                    if col_idx > len(row): continue
                    
                    # col_idx is 1-based from earlier logic? 
                    # If I manually set 5, it means 5th column. row tuple is 0-indexed.
                    # so row[col_idx-1]
                    cell_val = row[col_idx-1]
                    if not cell_val: continue
                    val_str = str(cell_val).strip()
                    if val_str.lower() in ['none', 'nan', '']: continue
                    
                    # Logic for "Claire (Tue) + Faye (Mon)" type cells
                    # If the cell contains brackets with day names, we filter.
                    # Otherwise we assume it applies to THIS day column (if in a grid)
                    # OR if it's a merged cell spanning all, we check days?
                    # The user snippet showed "Claire (Tuesday...)" inside a cell.
                    
                    target_staff_names = []
                    
                    # Split by newlines, "+", "&"
                    import re
                    # specific cleanup for "Claire (Tuesday...)" pattern
                    # If specific days are mentioned in brackets, only assign if matches current 'day'
                    
                    if "(" in val_str and any(d[:3].lower() in val_str.lower() for d in days_list):
                        # Complex cell
                        # Check if CURRENT day is mentioned
                        if day.lower() in val_str.lower() or day[:3].lower() in val_str.lower():
                            # Who is associated with this day?
                            # Regex: Name \((...)\)
                            # This is hard to parse perfectly without more robust NLP, but let's try strict tokenizing
                            # "Claire (Tuesday, Wednesday)" -> if today is Tuesday, add Claire.
                            # We can just look for the name preceding the bracket containing the day?
                            # Or simpler: Just look for names in the string.
                            # If a name is found, blindly assign it? NO, wrong day.
                            
                            # Heuristic: Split by "+" or newline.
                            parts = re.split(r'[+\n]', val_str)
                            for part in parts:
                                part = part.strip()
                                if not part: continue
                                
                                part_lower = part.lower()
                                # Check if this part has ANY day mentioned
                                # (Use full day names or 3-letter abbr)
                                days_mentioned_in_part = False
                                for d in days_list:
                                    if d.lower() in part_lower or d[:3].lower() in part_lower:
                                        days_mentioned_in_part = True
                                        break
                                
                                # Condition to assign:
                                # 1. The current day is explicitly mentioned in this part
                                # 2. OR No days are mentioned in this part (implies generic assignment for this cell's column)
                                current_day_match = (day.lower() in part_lower) or (day[:3].lower() in part_lower)
                                
                                if current_day_match or not days_mentioned_in_part:
                                    # Extract name from this part (remove brackets and content)
                                    name_part = part.split('(')[0].strip()
                                    if name_part:
                                        target_staff_names.append(name_part)

                    else:
                        # Simple cell (just names)
                        # Split by space if multiple? OR assuming one per line?
                        # User snippet: "K.Soe"
                        # But could be "K.Soe Baitoey"
                        names = re.split(r'[+\n&,]', val_str)
                        for n in names:
                            n = n.strip()
                            if n: target_staff_names.append(n)
                    
                    # Register Staff and Assignments
                    for raw_name in target_staff_names:
                        s_name = raw_name.strip()
                        # Clean up punctuation
                        s_name = s_name.replace(".", "") # "Nick.C" -> "NickC"? No "Nick.C" is the name.
                        # Wait, "Nick.C" is valid. "K.Soe".
                        # Maybe just remove trailing/leading non-alphanumeric?
                        
                        # Mapping check
                        map_name = s_name
                        if s_name.lower() == "nick": map_name = "Nick.C"
                        if s_name.lower() == "darryl": map_name = "Daryl" # typo fix
                        if s_name.lower() == "ginny": map_name = "Jinny"
                        
                        # Find or Create
                        staff_obj = db.query(Staff).filter(func.lower(Staff.name) == map_name.lower()).first()
                        
                        if not staff_obj:
                             # New staff found in Duty
                             is_qts = map_name in ["Mr Ben", "Nick.C"]
                             staff_obj = Staff(
                                 name=map_name,
                                 role="Teacher" if is_qts else "Duties Only",
                                 profile="Added from Duty Rota",
                                 is_active=True,
                                 is_specialist=False,
                                 can_cover_periods=is_qts # False for most duties
                             )
                             db.add(staff_obj)
                             db.flush() # get ID
                             print(f"    New Staff from Duty: {map_name} (Teacher={is_qts})")
                        
                        # Add Schedule
                        db.add(Schedule(
                            staff_id=staff_obj.id,
                            day_of_week=day,
                            period=period_num,
                            activity=f"{duty_name}: {desc_val}",
                            is_free=False
                        ))
        
        db.commit()

        # Process CCA Sheet
        if "CCA" in wb.sheetnames:
            print("--- Processing CCA ---")
            sheet = wb["CCA"]
            
            # Find Day Columns (standard header scan)
            day_cols = {}
            header_row_idx = None
            
            for r_idx, row in enumerate(sheet.iter_rows(max_row=20, values_only=True)):
                row_str = [str(c).lower() for c in row if c]
                if "monday" in row_str:
                    header_row_idx = r_idx + 1
                    for c_idx, val in enumerate(row):
                        if not val: continue
                        val_str = str(val).strip().lower()
                        for d in days_list:
                            if d.lower() in val_str:
                                day_cols[d] = c_idx + 1
                    break
            
            # If no header found, make a best guess or skip
            if not day_cols:
                 print("  Could not find day headers in CCA tab.")
            else:
                 print(f"  CCA Day Columns: {day_cols}")
                 
                 # Iterate CCA rows
                 for r_idx, row in enumerate(sheet.iter_rows(min_row=header_row_idx+1, values_only=True)):
                     if not row or all(c is None for c in row): continue
                     
                     # Assuming Activity Name is in Col A or B?
                     # Let's assume Col A (idx 0) is the Activity Name based on "Drone Club" example
                     cca_name = str(row[0]).strip() if row[0] else "CCA"
                     if not cca_name: cca_name = "CCA"
                     
                     # Iterate Day Columns
                     for day, col_idx in day_cols.items():
                         if col_idx > len(row): continue
                         cell_val = row[col_idx-1]
                         if not cell_val: continue
                         
                         val_str = str(cell_val).strip()
                         if val_str.lower() in ['none', 'nan', '']: continue
                         
                         # Split by separators: "+", "=", "\n"
                         # User mentioned "=" for Glee Club
                         import re
                         raw_names = re.split(r'[+=\n]', val_str)
                         
                         for raw_n in raw_names:
                             n = raw_n.strip()
                             if not n: continue
                             
                             # FILTER 1: Ignore Secondary staff (Sec)
                             if "(sec)" in n.lower():
                                 continue
                             
                             # CLEAN: Remove "K." or "Kun" prefix
                             # Regex: start of string, optional "k" or "kun" followed by dot or space
                             clean_name = re.sub(r'^(k\.|kun\s|k\s)', '', n, flags=re.IGNORECASE).strip()
                             
                             # Also remove specific day markers if inside the string e.g. "Claire (Tue)" -> "Claire"
                             # (Reusing logic from Duties might be safer, but let's just strip brackets for now)
                             clean_name = clean_name.split('(')[0].strip()
                             
                             # MAPPING
                             if clean_name.lower() == "nick": clean_name = "Nick.C"
                             if clean_name.lower() == "darryl": clean_name = "Daryl"
                             
                             # FILTER 2: STRICTLY EXISTING STAFF ONLY
                             staff_obj = db.query(Staff).filter(func.lower(Staff.name) == clean_name.lower()).first()
                             
                             if staff_obj:
                                 # Found existing staff
                                 # Assign CCA Period (Period 13 to separate from Duties/Lessons?)
                                 # Or just append to "Duty" checks? 
                                 # User said "end of the day". Let's use Period 13.
                                 # We need to make sure this period is checked during availability.
                                 
                                 db.add(Schedule(
                                     staff_id=staff_obj.id,
                                     day_of_week=day,
                                     period=13, # CCA Period
                                     activity=f"CCA: {cca_name}",
                                     is_free=False
                                 ))
                                 print(f"  Assigned CCA '{cca_name}' to {staff_obj.name} on {day}")
                             else:
                                 # Ignore new staff ("Ben P", "Gideon" etc)
                                 pass

        db.commit()
        print("Normalization complete.")
        
    except Exception as e:
        print(f"Error during normalization: {e}")
        db.rollback()
    finally:
        db.close()

if __name__ == "__main__":
    # Ensure tables exist
    Base.metadata.create_all(bind=engine)
    normalize_data()
