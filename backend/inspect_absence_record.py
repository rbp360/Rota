import openpyxl
import os

EXCEL_PATH = r"c:\Users\rob_b\Rota\temp_rota.xlsx"

def inspect_absence_record():
    if not os.path.exists(EXCEL_PATH):
        print(f"File not found: {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    sheet_name = 'Absence Record'
    if sheet_name not in wb.sheetnames:
        # Try case-insensitive or partial match
        matches = [s for s in wb.sheetnames if 'absence' in s.lower()]
        if matches:
            sheet_name = matches[0]
            print(f"Using sheet: {sheet_name}")
        else:
            print(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
            return

    sheet = wb[sheet_name]
    print(f"--- Inspecting {sheet_name} ---")
    for r_idx, row in enumerate(sheet.iter_rows(max_row=30, values_only=True)):
        print(f"Row {r_idx}: {row}")

if __name__ == "__main__":
    inspect_absence_record()
