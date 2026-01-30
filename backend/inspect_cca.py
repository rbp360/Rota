
import openpyxl
import pandas as pd

EXCEL_PATH = r"c:\Users\rob_b\Rota\temp_rota.xlsx"

def inspect_cca():
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        if "CCA" in wb.sheetnames:
            print("\n--- Inspecting CCA ---")
            sheet = wb["CCA"]
            for i, row in enumerate(sheet.iter_rows(max_row=30, values_only=True)):
                print(f"Row {i+1}: {row}")
        else:
            print("CCA sheet not found")
    except Exception as e:
        print(f"Error: {e}")


if __name__ == "__main__":
    import sys
    sys.stdout = open(r"c:\Users\rob_b\Rota\backend\cca_output.txt", "w")
    inspect_cca()

