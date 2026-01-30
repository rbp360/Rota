
import openpyxl

EXCEL_PATH = r"c:\Users\rob_b\Rota\temp_rota.xlsx"
wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
print("Sheet names:", wb.sheetnames)

if "CCA" in wb.sheetnames:
    sheet = wb["CCA"]
    print("\nCCA Sheet Content (First 15 rows):")
    for row in sheet.iter_rows(max_row=15, values_only=True):
        print(row)
