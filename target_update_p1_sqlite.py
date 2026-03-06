import os
import re
import openpyxl
import sqlite3

# Configuration
EXCEL_PATH = "GV cover and staff absence.xlsx"
DB_PATH = "data_archive/rota.db" # Targeted the archive one as it's the source for sync
DAYS_LIST = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

# Specialist list
SPECIALISTS = [
    "Daryl", "Becky", "Billy", "Jinny", "Ginny", "Ben", "Faye", 
    "Claire", "Jake", "Retno", "Jacinta", "Sunny", "Mr Ben", "Nick C"
]
FREE_SUBJECTS = ['thai', 'music', 'pe', 'p.e.', 'phse']

def clean_name(name):
    if not name: return ""
    n = str(name).strip()
    n = re.sub(r'^(mr|mrs|ms|miss|k\.|kun\s|k\s)\s*', '', n, flags=re.IGNORECASE).strip()
    n = n.split('(')[0].strip()
    nl = n.lower()
    if "jinny" in nl or "ginny" in nl: return "Jinny"
    return n

def target_update_p1_sqlite():
    if not os.path.exists(DB_PATH):
        print(f"Error: DB not found at {DB_PATH}")
        return

    print(f"Loading Excel: {EXCEL_PATH}...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    sheet_names = wb.sheetnames
    
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
    cursor.execute("SELECT id, name FROM staff")
    staff_mems = cursor.fetchall()
    
    update_count = 0
    
    for s in staff_mems:
        staff_id = s['id']
        original_name = s['name']
        staff_name = clean_name(original_name)
        
        if not staff_name: continue
        
        target_sheet = None
        for sname in sheet_names:
            if staff_name.lower() in sname.lower() or sname.lower() in staff_name.lower():
                target_sheet = sname
                break
        
        if not target_sheet:
            if staff_name == "Claire" and "ME" in sheet_names:
                target_sheet = "ME"
            else:
                continue

        print(f"Updating P1 for: {original_name}...")
        sheet = wb[target_sheet]
        
        day_cols = {}
        header_found = False
        for row in sheet.iter_rows(max_row=10, values_only=True):
            for i, cell in enumerate(row):
                if cell:
                    c_val = str(cell).strip().lower()
                    for d in DAYS_LIST:
                        if d.lower() in c_val:
                            day_cols[d] = i
                            header_found = True
            if header_found: break
        
        if not day_cols:
            day_cols = {d: i+4 for i, d in enumerate(DAYS_LIST)}

        p1_row_data = None
        for row in sheet.iter_rows(max_row=60, values_only=True):
            for cell in row[:5]:
                if cell:
                    c_val = str(cell).strip().lower()
                    if re.search(r'(^|\b)(period\s*|p\.?\s*|)1(\b|$)', c_val):
                        p1_row_data = row
                        break
            if p1_row_data: break
        
        if p1_row_data:
            for day, col_idx in day_cols.items():
                if col_idx < len(p1_row_data):
                    activity = str(p1_row_data[col_idx]).strip() if p1_row_data[col_idx] else ""
                    
                    is_free = 0 # SQLite uses 0/1
                    clean_act = activity.lower()
                    
                    if not clean_act or any(x in clean_act for x in ['free', 'available', 'none', 'nan']):
                        is_free = 1
                    elif staff_name not in SPECIALISTS:
                        if any(sub in clean_act for sub in FREE_SUBJECTS):
                            is_free = 1
                    
                    # Update local DB
                    cursor.execute("""
                        UPDATE schedules 
                        SET activity = ?, is_free = ? 
                        WHERE staff_id = ? AND day_of_week = ? AND period = 1
                    """, (activity, is_free, staff_id, day))
                    
                    # If no row updated, it might be a new entry
                    if cursor.rowcount == 0:
                        cursor.execute("""
                            INSERT INTO schedules (staff_id, day_of_week, period, activity, is_free)
                            VALUES (?, ?, 1, ?, ?)
                        """, (staff_id, day, activity, is_free))
                    
                    update_count += 1
            
    conn.commit()
    conn.close()
    print(f"\n✅ LOCAL UPDATE FINISHED. {update_count} records updated in SQLite.")
    print("Next step is to sync this to Cloud.")

if __name__ == "__main__":
    target_update_p1_sqlite()
