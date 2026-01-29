import openpyxl
import os

file_path = r"c:\Users\rob_b\Rota\GV cover and staff absence.xlsx"

def inspect_layout():
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        # Try to find a teacher sheet (Daryl or Faye or Jacinta)
        sheet_name = "Daryl" if "Daryl" in wb.sheetnames else wb.sheetnames[1]
        sheet = wb[sheet_name]
        
        with open("layout_debug.txt", "w") as f:
            f.write(f"Inspecting Sheet: {sheet_name}\n")
            for r_idx, row in enumerate(sheet.iter_rows(max_row=20, max_col=15, values_only=True)):
                f.write(f"Row {r_idx}: {row}\n")
        print("Layout debug saved to layout_debug.txt")
    except Exception as e:
        print(f"Error: {e}")

if __name__ == "__main__":
    inspect_layout()
