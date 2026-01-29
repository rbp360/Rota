import openpyxl
import os

file_path = r"c:\Users\rob_b\Rota\GV cover and staff absence.xlsx"

def dump_all():
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        print(f"SHEETS|{','.join(wb.sheetnames)}")
        for name in wb.sheetnames:
            print(f"SHEET_START|{name}")
            sheet = wb[name]
            for row in sheet.iter_rows(max_row=10, max_col=10, values_only=True):
                print(f"ROW|{row}")
            print(f"SHEET_END|{name}")
    except Exception as e:
        print(f"ERROR|{e}")

if __name__ == "__main__":
    dump_all()
