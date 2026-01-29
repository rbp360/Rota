import openpyxl

EXCEL_PATH = r"c:\Users\rob_b\Rota\GV cover and staff absence.xlsx"
days_list = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

def debug_jill():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    if "Jill" not in wb.sheetnames:
        print("Jill sheet not found")
        return
    
    sheet = wb["Jill"]
    print(f"--- Jill Sheet Layout ---")
    day_cols = {}
    header_row_idx = 0
    for r_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
        print(f"Row {r_idx + 1}: {row}")
        found_in_row = False
        for idx, val in enumerate(row):
            if val and str(val).strip() in days_list:
                day_cols[str(val).strip()] = idx + 1
                found_in_row = True
        if found_in_row:
            header_row_idx = r_idx + 1
            break
            
    print(f"Found Days at Row: {header_row_idx}")
    print(f"Day Cols: {day_cols}")
    
    rows = list(sheet.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 10, values_only=True))
    for p_idx in range(1, 7):
        if p_idx > len(rows): continue
        row_data = rows[p_idx-1]
        print(f"P{p_idx} Data: {row_data}")
        for day, col in day_cols.items():
            raw_val = row_data[col-1] if col <= len(row_data) else None
            val = str(raw_val).strip() if raw_val is not None else ""
            is_free = not val or val.lower() in ['none', 'nan', 'free', 'available']
            print(f"  {day} P{p_idx}: '{val}' (Free: {is_free})")

if __name__ == "__main__":
    debug_jill()
