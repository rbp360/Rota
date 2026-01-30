
import openpyxl
import os

def find_cell(query):
    wb = openpyxl.load_workbook(r"c:\Users\rob_b\Rota\temp_rota.xlsx", data_only=True)
    sheet = wb["CCA"]
    for r_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        for c_idx, val in enumerate(row):
            if val and query.lower() in str(val).lower():
                print(f"Found '{query}' at R{r_idx+1} C{c_idx+1}: {val}")
    
    # Also print some context around it
    print("\n--- Context (Row 1 headers) ---")
    headers = next(sheet.iter_rows(max_row=1, values_only=True))
    print(headers)

if __name__ == "__main__":
    find_cell("Eco action")
    find_cell("Rosie")
    find_cell("Thursday")
